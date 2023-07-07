# Small utility to ask user for an input file 
# containing a list of values (in this example it is a text string)
# and parse a spreadsheet for the existance of a value within the entire woorkbook

import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import ttk
from openpyxl import load_workbook

# Function to handle the search and generate the output
def search_in_workbook():
    # Ask user to select the input file
    input_file_path = filedialog.askopenfilename(title="Select Input File", filetypes=(("Text Files", "*.txt"),))

    if input_file_path:
        # Ask user to select the workbook file
        workbook_path = filedialog.askopenfilename(title="Select Workbook", filetypes=(("Excel Files", "*.xlsx"),))

        if workbook_path:
            # Load the workbook
            wb = load_workbook(workbook_path)

            # Get all sheet names from the workbook
            sheet_names = wb.sheetnames

            # Read each line from the input file
            not_found_values = []
            with open(input_file_path, 'r') as input_file:
                for line in input_file:
                    search_value = line.strip()

                    # Search for the value in each sheet of the workbook
                    found = False
                    for sheet_name in sheet_names:
                        sheet = wb[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                if cell.value == search_value:
                                    found = True
                                    break
                            if found:
                                break
                        if found:
                            break

                    # Add the value to the not found list if not found in any sheet
                    if not found:
                        not_found_values.append(search_value)

            # Close the workbook
            wb.close()

            # Generate the output message
            if not_found_values:
                output_message = "Values not found:\n\n" + "\n".join(not_found_values)
            else:
                output_message = "All values found!"

            # Create the root window
            root = tk.Tk()
            root.withdraw()

            # Create a frame for the text area and scrollbar
            frame = ttk.Frame(root)
            frame.pack(fill=tk.BOTH, expand=True)

            # Create a vertical scrollbar
            scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL)

            # Create a text area to display the results
            output_text = tk.Text(frame, wrap=tk.WORD, yscrollcommand=scrollbar.set)
            output_text.insert(tk.END, output_message)
            output_text.configure(state='disabled')

            # Configure the scrollbar
            scrollbar.config(command=output_text.yview)

            # Pack the text area and scrollbar
            output_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            result = messagebox.askyesno("Search Result", "", icon='info', detail=output_message, parent=root)

            # Handle the button clicks
            if result:
                # Ask user to select the save location
                save_file_path = filedialog.asksaveasfilename(defaultextension='.txt', filetypes=(("Text Files", "*.txt"),))
                if save_file_path:
                    with open(save_file_path, 'w') as save_file:
                        save_file.write("\n".join(not_found_values))
                    messagebox.showinfo("Save Successful", "File saved successfully!")

# Call the search function
search_in_workbook()
